from io import BytesIO
from openpyxl import load_workbook
from django.db import transaction
from products.models import Product
from products.serializers import ProductCreateSerializer

def parse_null_values(value):
    return None if value in [None, 'Null', 0] else  'None'

class ExcelProductParser:
    """
    Parses an Excel (.xlsx) file and returns a list of product dictionaries
    ready to be validated by a DRF serializer.
    """

    def parse(self, file) -> list[dict]:
        """
        file: InMemoryUploadedFile | TemporaryUploadedFile
        """
        workbook = load_workbook(
            filename=BytesIO(file.read()), read_only=True, data_only=True
        )
        sheet = workbook.active

        products: list[dict] = []

        for index, row in enumerate(
            sheet.iter_rows(min_row=2, values_only=True), start=2
        ):
            # Skip empty rows
            if not row or not row[0]:
                continue
            product = {
                "sku": str(row[0]).strip(),
                "name": row[1],
                "description": row[2],
                "price": float(row[3]) if row[3] is not None else 0,
                "discount_price": float(row[4]) if row[4] else None,
                "purchase_price": float(row[5]) if row[5] is not None else 0,
                "stock": int(row[6]) if row[6] is not None else 0,
                "category": 1,  # row[7],  # FK id
                "score": int(row[8]) if row[8] else None,
                "recommended": bool(row[9]) if row[9] is not None else False,
                "best_seller": bool(row[10]) if row[10] is not None else False,
                "tag": row[11] or "",
                "quality": row[12] or "",
                "weight": float(row[13]) if row[13] is not None else 0,
                "slug": row[14],
                "unit_of_measurement": None#parse_null_values(row[15])
            }

            products.append(product)

        workbook.close()
        return products


class ProductBulkCreateService:
    """
    Class performs a `Model.objects.bulk_create()` operation
    Raise an Exception if the operation can't be performanced
    """

    def execute(self, products_data: list[dict]) -> dict:
        serializer = ProductCreateSerializer(data=products_data, many=True)

        serializer.is_valid(raise_exception=True)

        products = [Product(**data) for data in serializer.validated_data]

        with transaction.atomic():
            Product.objects.bulk_create(products)

        return {"created": len(products)}

from datetime import timedelta
from io import BytesIO
import uuid
from django.db import transaction
from django.contrib.auth.hashers import make_password
from django.utils import timezone
# from openpyxl import load_workbook
from users.models import User
from openpyxl import load_workbook
from users.models import User, UserProfileSettings, ReferralDiscount
from users.serializers import BulkCreateUserSerializer
from users.utils.profile import create_user_profile_settings
from users.tasks import send_welcome_email


def set_discount_is_valid_until():
    return timezone.now() + timedelta(days=8)



class ExcelUserParser:
    """
    Parses an Excel (.xlsx) file and returns a list of product dictionaries
    ready to be validated by a DRF serializer.
    """

    def parse(self, file) -> list[dict]:
        """
        file: InMemoryUploadedFile | TemporaryUploadedFile
        """

        workbook = load_workbook(
            filename=BytesIO(file.read()), read_only=True, data_only=True
        )

        sheet = workbook.active
        users: list[dict] = []

        for idx, row in enumerate(
            sheet.iter_rows(min_row=2, values_only=True), start=2
        ):
            # Skip empty rows
            if not row or not row[0]:
                continue

            user = {
                "dni": str(int(row[0])).strip(),
                "first_name": row[1],
                "last_name": row[2],
                "phone": row[3],
                "email": row[4],
                "password": make_password(str(int(row[0])).strip()),
                "username": self.make_username(
                    str(row[0]).strip(), str(row[1]).strip()
                ),
                "referral_code": self.get_referral_code(),
                "active": True,
                "user_groups": [],
                "role": row[5] or "customer",
            }

            users.append(user)

        workbook.close()
        return users

    def make_username(self, dni: str, name: str) -> str:
        prefix_dni = str(dni[::-5])
        prefix_name = name
        username = prefix_name + prefix_dni

        return username.strip().lower()

    def get_referral_code(self) -> str:
        prefix: str = "AVB-RFC"
        return f"{prefix}-{str(uuid.uuid4().hex[:12])}".upper()


class UsersBulkCreate:
    """
    Class performs a `Model.objects.bulk_create()` operation
    Raise an Exception if the operation can't be performanced
    """

    @classmethod
    def execute(cls, users_data: list[dict]) -> dict:
        serializer = BulkCreateUserSerializer(data=users_data, many=True)

        serializer.is_valid(raise_exception=True)
        users = [User(**data) for data in serializer.validated_data]
        
        with transaction.atomic():
            users = User.objects.bulk_create(users)

            profiles = [
                UserProfileSettings(user=user)
                for user in users
            ]

            referral_discounts = [
                ReferralDiscount(
                    user=user,
                    has_discount=True,
                    expires_at=set_discount_is_valid_until()
                )
                for user in users
            ]


            UserProfileSettings.objects.bulk_create(profiles)
            ReferralDiscount.objects.bulk_create(referral_discounts)

            transaction.on_commit(lambda: [
                send_welcome_email.delay(user.email, user.username)
                for user in users
            ])
            
        return {"created": len(users)}
